
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
from odoo import api, exceptions ,models, fields, _
from odoo.exceptions import UserError
import numpy
from dateutil import rrule
from odoo import SUPERUSER_ID
from odoo.http import request
import requests
from odoo.tools.safe_eval import safe_eval, test_python_expr
import json
import base64
from io import StringIO
import logging
import pathlib
import os

import openpyxl
from openpyxl.styles import Border, Font, Alignment, numbers, NumberFormatDescriptor,Side, PatternFill
from openpyxl.styles import colors, borders, fills


class Notimove(models.Model):
    _inherit ="account.move"

    def ks_apply_left(self, ks_cell, kc='', vc='', sz=False, wp=False, fm=0):
        ks_cell.alignment = Alignment(horizontal="right" if kc else 'center', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=wp, size=sz)
        if fm == 1: ks_cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'
        if fm == 2: ks_cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        if fm == 3: ks_cell.number_format = '0%'


    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp, shrinkToFit=False)
        if sz: ks_cell.font = Font(b=True, size=sz)


    def Cxp_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = "Nombre"
        self.ks_apply_style(sheet['A1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['B1'] = "Descripcion"
        self.ks_apply_style(sheet['B1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['C1'] = "Cuenta"
        self.ks_apply_style(sheet['C1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['D1'] = "Fecha"
        self.ks_apply_style(sheet['D1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['E1'] = "Monto Residual"
        self.ks_apply_style(sheet['E1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['F1'] = "Monto Total"
        self.ks_apply_style(sheet['F1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

    def Cxc_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = "Nombre"
        self.ks_apply_style(sheet['A1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['B1'] = "Manzana"
        self.ks_apply_style(sheet['B1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['C1'] = "Lote"
        self.ks_apply_style(sheet['C1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['D1'] = "Contrato"
        self.ks_apply_style(sheet['D1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['E1'] = "Fecha de Vencimiento"
        self.ks_apply_style(sheet['E1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['F1'] = "Meses vencidos"
        self.ks_apply_style(sheet['F1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['G1'] = "Audeudo con mora"
        self.ks_apply_style(sheet['G1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['H1'] = "Saldo Insoluto"
        self.ks_apply_style(sheet['H1'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

    @api.model
    def generate_mail_cxp(self):
        busqueda = self.env['ji.notification.slow.payer'].search([('ji_models', '=', 'account.move')])
        for noti in busqueda:
            today = fields.Date.context_today(self)
            if noti.active and noti.type == 'cxp':


                for lines in noti.notification_lines:
                    factu = self.env['account.move'].search(
                            [('type', '=', 'in_invoice'), ('state', '=', 'posted'), ('invoice_payment_state','!=', 'paid')])
                    if lines.partner_id:
                        factu = self.env['account.move'].search(
                            [('type', '=', 'in_invoice'), ('state', '=', 'posted'),('partner_id','=',lines.partner_id.id),('invoice_payment_state','!=', 'paid')])

                    template_id = self.env['ir.model.data'].get_object_reference('jibaritolotes', 'mail_template_cuenta_cobrar')[1]
                    email_template_obj = self.env['mail.template'].browse(template_id)

                    con = False

                    if noti.recurrencia == 'semana':

                        con = today.strftime("%w") == lines.name

                    elif noti.recurrencia == 'mes':
                        if lines.name == today.day:
                            con = True
                    #raise UserError(_(f'DATOS {factu}'))
                    if template_id and con:
                        data = []
                        mail_from = ""
                        ids= 0
                        report_name = "Cuentas por Pagar"
                        workbook = openpyxl.Workbook()

                        sheet = workbook.active

                        self.Cxp_create_workbook_header(report_name, sheet)
                        i = 1
                        row = 2
                        col = 0
                        for account in factu:

                            mail_from = account.company_id.partner_id.email
                            ids = account.id
                            sheet.cell(row, 1, account.partner_id.name)

                            # linea facturable
                            desc = ""
                            accu = ""
                            for linea in account.invoice_line_ids:
                                desc = linea.name
                            accu = account.codigo_prod.name

                            sheet.cell(row, 2, desc)
                            sheet.cell(row, 3, accu)
                            sheet.cell(row, 4, account.date)
                            sheet.cell(row, 5, account.amount_residual)
                            sheet.cell(row, 6, account.amount_total)
                            row += 1
                            i += 1

                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter  # Get the column name
                            for cell in col:
                                try:  # Necessary to avoid error on empty cells
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column].width = adjusted_width
                        output = StringIO()
                        # tupla = numpy.asarray(data)
                        #raise UserError(_(f'DATOS {tupla}'))
                        # numpy.savetxt('/tmp/file.csv', tupla, delimiter=";", newline="\n", fmt="%s")
                        lectura = ""
                        # with open("/tmp/file.csv", "rb") as rfile:
                            #encoded = base64.b64encode(rfile.read())
                            #
                            # lectura= rfile.read()
                        # data_record = base64.encodebytes((lectura).encode())
                        filename = ('/tmp/' + str(report_name) + '.xlsx')
                        workbook.save(filename)
                        fp = open(filename, "rb")
                        file_data = fp.read()
                        data_record = base64.b64encode(file_data)
                        att = {
                                'name': str(report_name) + '.xlsx',
                                'type': 'binary',
                                'datas': data_record,
                                'store_fname': data_record,
                                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            }
                        data_id = self.env['ir.attachment'].create(att)



                        values = email_template_obj.generate_email(325, fields=None)
                        values['email_from'] = mail_from
                        # value['email_to'] = partner.email
                        values['email_to'] = str([partner.email for partner in noti.partner_ids]).replace('[',
                                                                                                          '').replace(
                            ']', '').replace("'", "")
                        values['subject'] = noti.object
                        # raise UserError(_(value['email_to']))
                        values['attachment_ids'] = [(6, 0, [data_id.id])]

                        # values['email_from'] = "jibarito2005@gmail.com"
                        # value['email_to'] = partner.email
                        # values['email_to'] = "anagrv1981@hotmail.com, jesus.nazareth@ogum.com.mx"
                        # values['subject'] = "Prueba de correo auto"

                        # values['res_id'] = True
                        # values['author_id'] = self.env['res.users'].browse(request.env.uid).partner_id.id
                        mail_mail_obj = self.env['mail.mail']
                        print('\n\n\n', values, '\n\n\n')
                        msg_id = mail_mail_obj.sudo().create(values)
                        if msg_id:
                            mail_mail_obj.sudo().send([msg_id])
    @api.model
    def prueba_whasthapp(self):
        url = "https://api.apichat.io/v1/sendText"
        headers = {
            "Content-Type": "application/json",
            "client-id": "21746",
            "token": "hnw36fDy44HL"
        }
        body = {"number": "5215581692962", "text": "Que onda aprin debes $45000!  🤝"}

        requests.post(url, json=body, headers=headers)
    @api.model
    def generate_mail_cxc(self):

        busqueda = self.env['ji.notification.slow.payer'].search([('ji_models', '=', 'account.move')])
        for noti in busqueda:

            today = fields.Date.context_today(self)
            if noti.active and noti.type == 'cxc':

                for lines in noti.notification_lines:
                    busqueda = [('type', '=', 'out_invoice'), ('state', '=', 'posted'),('invoice_payment_state', '!=', 'paid')]

                    if lines.partner_id:
                        busqueda.append(('partner_id', '=', lines.partner_id.id))
                    if noti.is_mora:
                        busqueda.append(('total_mes1', '>', 0))
                    factu = self.env['account.move'].search(busqueda)


                    template_id = self.env['ir.model.data'].get_object_reference('jibaritolotes', 'mail_template_cuenta_cobrar')[1]
                    email_template_obj = self.env['mail.template'].browse(template_id)

                    con = False

                    if noti.recurrencia == 'semana':

                        con = today.strftime("%w") == lines.name

                    elif noti.recurrencia == 'mes':
                        if lines.name == today.day:
                                con = True


                    if template_id and con:

                        data = []
                        # mail_from = ""
                        ids = 0
                        report_name = "Cuentas por Cobrar"
                        workbook = openpyxl.Workbook()

                        sheet = workbook.active

                        self.Cxc_create_workbook_header(report_name, sheet)
                        i = 1
                        row = 2
                        col = 0

                        for account in factu:
                            account.action_moratorio_v()
                            mail_from = account.company_id.partner_id.email
                            ids = account.id
                            sheet.cell(row, 1, account.partner_id.name)

                            # linea facturable
                            date_ultpay = fields.Date.from_string(account.proxfecha_venci)
                            total = account.amount_residual + account.total_moratorium

                            sheet.cell(row, 2, account.x_studio_manzana.name)
                            sheet.cell(row, 3, account.x_studio_lote.name)
                            sheet.cell(row, 4, account.name)
                            sheet.cell(row, 5, date_ultpay)
                            sheet.cell(row, 6, account.total_mes)
                            sheet.cell(row, 7, account.saldo_pend)
                            sheet.cell(row, 8, total)
                            row += 1
                            i += 1
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter  # Get the column name
                            for cell in col:
                                try:  # Necessary to avoid error on empty cells
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column].width = adjusted_width

                        output = StringIO()
                        # tupla = numpy.asarray(data)
                        # raise UserError(_(f'DATOS {tupla}'))
                        # numpy.savetxt('/tmp/file.csv', tupla, delimiter=";", newline="\n", fmt="%s")
                        lectura = ""
                        # with open("/tmp/file.csv", "rb") as rfile:
                        # encoded = base64.b64encode(rfile.read())
                        #
                        # lectura= rfile.read()
                        # data_record = base64.encodebytes((lectura).encode())
                        filename = ('/tmp/' + str(report_name) + '.xlsx')
                        workbook.save(filename)
                        fp = open(filename, "rb")
                        file_data = fp.read()
                        data_record = base64.b64encode(file_data)
                        att = {
                            'name': str(report_name) + '.xlsx',
                            'type': 'binary',
                            'datas': data_record,
                            'store_fname': data_record,
                            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        }
                        data_id = self.env['ir.attachment'].create(att)

                        values = email_template_obj.generate_email(325, fields=None)
                        values['email_from'] = mail_from
                        # value['email_to'] = partner.email
                        values['email_to'] = str([partner.email for partner in noti.partner_ids]).replace('[','').replace(']', '').replace("'", "")
                        values['subject'] = noti.object
                        # raise UserError(_(value['email_to']))
                        values['attachment_ids'] = [(6, 0, [data_id.id])]



                        # values['email_from'] = "jibarito2005@gmail.com"
                        # value['email_to'] = partner.email
                        # values['email_to'] = "anagrv1981@hotmail.com, jesus.nazareth@ogum.com.mx"
                        # values['subject'] = "Prueba de correo auto"

                        # values['res_id'] = True
                        # values['author_id'] = self.env['res.users'].browse(request.env.uid).partner_id.id
                        mail_mail_obj = self.env['mail.mail']
                        print('\n\n\n', values, '\n\n\n')
                        msg_id = mail_mail_obj.sudo().create(values)
                        if msg_id:
                            mail_mail_obj.sudo().send([msg_id])

    @api.model
    def genera_report_gastos(self):
        busqueda = self.env['ji.notification.slow.payer'].search([('ji_models', '=', 'account.move')])
        for noti in busqueda:
            today = fields.Date.context_today(self)
            if noti.active and noti.type == 'financiero':
                con = False
                for lines in noti.notification_lines:
                    if lines.name == today.day:
                        con = True
                if con:
                    ms = datetime.strptime(str(today.strftime("%Y-%m-")) + "01", "%Y-%m-%d")
                    es_mes = fields.Date.from_string(ms)
                    pas_mes = fields.Date.from_string(ms)
                    # date_ultpay -= relativedelta(months=account.total_mes)


                    # if noti.recurrencia == 'semana':
                    #
                    #     con = today.strftime("%w")

                    # elif noti.recurrencia == 'mes':

                    pas_mes -= relativedelta(months=1)
                    es_mes -= relativedelta(days=1)

                    categoria = self.env['product.category'].search([('name', '=', 'Vista Bella')])
                    document = self.set_report_gasto(categoria,pas_mes,es_mes)

                    template_id = \
                    self.env['ir.model.data'].get_object_reference('jibaritolotes', 'mail_template_reporte_gastos')[1]
                    email_template_obj = self.env['mail.template'].browse(template_id)
                    if template_id:
                        data_record = base64.b64encode(document[0])
                        att = {
                            'name': 'Reporte de Gastos.pdf',
                            'type': 'binary',
                            'datas': data_record,
                            'store_fname': data_record,
                            'mimetype': 'application/x-pdf'
                        }
                        data_id = self.env['ir.attachment'].create(att)

                        values = email_template_obj.generate_email(325, fields=None)
                        values['email_from'] = "jibarito2005@gmail.com"
                        # value['email_to'] = partner.email
                        values['email_to'] = str([partner.email for partner in noti.partner_ids]).replace('[', '').replace(
                            ']', '').replace("'", "")
                        values['subject'] = noti.object
                        # raise UserError(_(value['email_to']))
                        values['attachment_ids'] = [(6, 0, [data_id.id])]

                        mail_mail_obj = self.env['mail.mail']
                        print('\n\n\n', values, '\n\n\n')
                        msg_id = mail_mail_obj.sudo().create(values)
                        if msg_id:
                            mail_mail_obj.sudo().send([msg_id])
                #

    @api.model
    def generate_mail_move(self):
        #su_id = self.env['account.move'].browse(SUPERUSER_ID)

        busqueda = self.env['ji.notification.slow.payer'].search([('ji_models', '=', 'account.move')])
        for noti in busqueda:
            today = fields.Date.context_today(self)
            if noti.active and noti.type == 'cliente':
            
                for lines in noti.notification_lines:
                    factu = self.env['account.move'].search(
                            [('type', '=', 'out_invoice'), ('state', '=', 'posted'), ])
                    if lines.partner_id:
                        factu = self.env['account.move'].search(
                            [('type', '=', 'out_invoice'), ('state', '=', 'posted'),('partner_id','=',lines.partner_id.id), ])
                    for account in factu:
                        partner = account.partner_id

                        date_ultpay = fields.Date.from_string(account.proxfecha_venci)
                        # date_ultpay -= relativedelta(months=account.total_mes)
                        con = False

                        if noti.recurrencia == 'semana':
                                date_ultpay = date_ultpay
                                con = today.strftime("%w") == lines.name

                        elif noti.recurrencia == 'mes':
                            if lines.is_mora:
                                # date_ultpay -= relativedelta(months=account.total_mes)
                                # date_ultpay += relativedelta(months=lines.name)
                                # date_ultpay = date_ultpay.date()
                                if lines.name == today.day:
                                    con = True
                            else:
                                date_ultpay -= relativedelta(months=lines.name)
                                date_ultpay = date_ultpay
                                if lines.name == today.day and date_ultpay.month == today.month:
                                    con = True


                        elif noti.recurrencia == 'periodo':
                            if lines.is_mora:
                                date_ultpay += relativedelta(day=lines.name)
                                date_ultpay = date_ultpay.date()
                                con = date_ultpay == today
                            else:
                                date_ultpay -= relativedelta(day=lines.name)
                                date_ultpay = date_ultpay.date()
                                con = date_ultpay == today

                        #
                        template_id = self.env['ir.model.data'].get_object_reference('jibaritolotes','mail_template_reporte_gastos')[1]
                        #
                        email_object = self.env['mail.template'].browse(template_id)

                        if template_id and con:
                            value = email_object.generate_email(account.id)
                            value['email_from'] = account.company_id.partner_id.email
                            value['email_to'] = partner.email
                            # value['partner_to'] = str([partner.id for partner in noti.partner_ids]).replace('[', '').replace(']', '')
                            value['subject'] = noti.object
                            # value['model_id'] = 'account.move'
                            value['author_id'] = self.env['res.users'].browse(request.env.uid).partner_id.id
                            mail_mail_obj = self.env['mail.mail']
                            # raise UserError(_(str([partner.id for partner in noti.partner_ids]).replace('[', '').replace(']', '')))
                            msg_id = mail_mail_obj.sudo().create(value)
                            if msg_id:
                                mail_mail_obj.sudo().send([msg_id])


    @api.model
    def generate_email_etado_cuenta(self):
        # self.prueba_whasthapp()
        busqueda = self.env['ji.notification.slow.payer'].search([('ji_models', '=', 'account.move')])
        for noti in busqueda:
            today = fields.Date.context_today(self)
            if noti.active and noti.type == 'clien_estado':

                for lines in noti.notification_lines:
                    busqueda = [('type', '=', 'out_invoice'), ('state', '=', 'posted'), ]

                    if lines.partner_id:
                            busqueda.append(('partner_id', '=', lines.partner_id.id))
                    if noti.is_mora:
                        busqueda.append(('ji_is_moratorium', '=', noti.is_mora))
                    factu = self.env['account.move'].search(busqueda)
                    for account in factu:
                        partner = account.partner_id

                        date_ultpay = fields.Date.from_string(account.proxfecha_venci)
                        # date_ultpay -= relativedelta(months=account.total_mes)
                        con = False

                        if noti.recurrencia == 'semana':
                            date_ultpay = date_ultpay
                            if today.strftime("%w") == lines.name:
                                con = True
                        elif noti.recurrencia == 'mes':
                                if lines.name == today.day:
                                    con = True
                        elif noti.recurrencia == 'periodo':

                                date_ultpay -= relativedelta(months=1)
                                date_ultpay += timedelta(days=1)
                                if date_ultpay == today:
                                    con = True

                        #
                        template_id = self.env['ir.model.data'].get_object_reference('jibaritolotes', 'mail_template_estado_cuenta')[1]
                        #
                        email_object = self.env['mail.template'].browse(template_id)
                        # raise UserError(_(str(con)))
                        if template_id and con:

                            document = factu.get_reporte_amoritizacionv2_pdf(account)

                            data_record = base64.b64encode(document[0])
                            ir_values = {
                                'name': "Estado de Cuenta",
                                'type': 'binary',
                                'datas': data_record,
                                'store_fname': data_record,
                                'mimetype': 'application/x-pdf',
                            }
                            data_id = self.env['ir.attachment'].create(ir_values)
                            email_object.partner_to = str([partner.id for partner in noti.partner_ids]).replace('[', '').replace(']', '').replace("'","")
                            value = email_object.generate_email(account.id)
                            value['email_from'] = account.company_id.partner_id.email
                            value['email_to'] = partner.email
                            value['recipient_ids'] = noti.partner_ids
                            value['subject'] = noti.object
                            # value['model_id'] = 'account.move'
                            value['attachment_ids'] = [(6, 0, [data_id.id])]
                            value['author_id'] = self.env['res.users'].browse(request.env.uid).partner_id.id
                            mail_mail_obj = self.env['mail.mail']
                            # raise UserError(_(str([partner.id for partner in noti.partner_ids]).replace('[', '').replace(']', '')))
                            msg_id = mail_mail_obj.sudo().create(value)
                            if msg_id:
                                mail_mail_obj.sudo().send([msg_id])
    def border_range(self,range, border):
        for cell in range:
            for x in cell:
                x.border = border
    def fill_range(self,range, fill):
        for cell in range:
            for x in cell:
                x.fill = fill
    def docenal_create_workbook_header(self, report_name, sheet, dia):
        sheet.title = str(report_name)

        thin_border = Border(left=Side(border_style=borders.BORDER_THIN, color=colors.Color('ED7D31')),
                             right=Side(border_style=borders.BORDER_THIN, color=colors.Color('ED7D31')),
                             top=Side(border_style=borders.BORDER_THIN, color=colors.Color('ED7D31')),
                             bottom=Side(border_style=borders.BORDER_THIN, color=colors.Color('ED7D31')))

        self.border_range(sheet['B6:F11'],thin_border)
        self.border_range(sheet['B13:F16'],thin_border)
        self.border_range(sheet['B18:F23'],thin_border)


        thin_fill = PatternFill(start_color="F9DA78", end_color="F9DA78", fill_type="solid")
        self.fill_range(sheet['B6:F6'],thin_fill)
        self.fill_range(sheet['B13:F13'],thin_fill)
        self.fill_range(sheet['B18:F18'],thin_fill)

        thin_fill = PatternFill(start_color="FEF2CC", end_color="FEF2CC", fill_type="solid")
        self.fill_range(sheet['D7:D11'],thin_fill)
        self.fill_range(sheet['D14:D16'],thin_fill)
        self.fill_range(sheet['D19:D23'],thin_fill)
        self.fill_range(sheet['F7:F11'],thin_fill)
        self.fill_range(sheet['F14:F16'],thin_fill)
        self.fill_range(sheet['F19:F23'],thin_fill)

        sheet['B3'] = "Indicadores Vista Bella El Jibarito"
        self.ks_apply_style(sheet['B3'], True, True, 16, True)
        sheet.merge_cells('B3:F3')

        sheet['B4'] = dia
        self.ks_apply_style(sheet['B4'], True, True, 14, True)
        sheet.merge_cells('B4:F4')

        # fila 1
        sheet['B6'] = "Tipo de ingreso"
        self.ks_apply_style(sheet['B6'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['B7'] = "Enganches"
        self.ks_apply_left(sheet['B7'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B8'] = "Mensualidades"
        self.ks_apply_left(sheet['B8'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B9'] = "Mensualidades WF"
        self.ks_apply_left(sheet['B9'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B10'] = "Devoluciones a Clientes"
        self.ks_apply_left(sheet['B10'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B11'] = "Otros Productos"
        self.ks_apply_left(sheet['B11'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['C6'] = "Mes"
        self.ks_apply_style(sheet['C6'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['D6'] = "Decena 1"
        self.ks_apply_style(sheet['D6'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['E6'] = "Decena 2"
        self.ks_apply_style(sheet['E6'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['F6'] = "Decena 3"
        self.ks_apply_style(sheet['F6'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        # fila 2
        sheet['B13'] = "Concepto"
        self.ks_apply_style(sheet['B13'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B14'] = "Gastos del periodo"
        self.ks_apply_left(sheet['B14'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B15'] = "Cuentas por pagar"
        self.ks_apply_left(sheet['B15'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B16'] = "Cuentas por cobrar"
        self.ks_apply_left(sheet['B16'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['C13'] = "Mes"
        self.ks_apply_style(sheet['C13'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['D13'] = "Decena 1"
        self.ks_apply_style(sheet['D13'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['E13'] = "Decena 2"
        self.ks_apply_style(sheet['E13'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['F13'] = "Decena 3"
        self.ks_apply_style(sheet['F13'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        # fila 3
        sheet['B18'] = "Detalle de Cartera"
        self.ks_apply_style(sheet['B18'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B19'] = "Total"
        self.ks_apply_left(sheet['B19'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B20'] = "Al día"
        self.ks_apply_left(sheet['B20'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B21'] = "Vencidos 1 mensualidad"
        self.ks_apply_left(sheet['B21'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B22'] = "Vencidos 2 mensualidades"
        self.ks_apply_left(sheet['B22'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)
        sheet['B23'] = "Vencidos 3 o más mensualidades"
        self.ks_apply_left(sheet['B23'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['C18'] = "Numero Contratos"
        self.ks_apply_style(sheet['C18'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['D18'] = "Porcentaje"
        self.ks_apply_style(sheet['D18'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['E18'] = "Importe"
        self.ks_apply_style(sheet['E18'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        sheet['F18'] = "Interés"
        self.ks_apply_style(sheet['F18'], True, True, 12, True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

    def mes_c(self,mes):
        if str(mes) == "01":
            return "Enero"
        if str(mes) == "02":
            return "Febrero"
        if str(mes) == "03":
            return "Marzo"
        if str(mes) == "04":
            return "Abril"
        if str(mes) == "05":
            return "Mayo"
        if str(mes) == "06":
            return "Junio"
        if str(mes) == "07":
            return "Julio"
        if str(mes) == "08":
            return "Agosto"
        if str(mes) == "09":
            return "Septimbre"
        if str(mes) == "10":
            return "Octubre"
        if str(mes) == "11":
            return "Noviembre"
        if str(mes) == "12":
            return "Diciembre"
        if str(mes) == "":
            return ""
    def rep_docenal_email(self):
        today = fields.Date.context_today(self)
        # today = today.strftime("%Y-%m-%d")
        last_mes = fields.Date.from_string(today)

        Categorias = self.env['product.category'].search([('name', '=', 'Vista Bella')])

        if str(today.strftime("%d")) == "11" or str(today.strftime("%d")) == "21" or str(today.strftime("%d")) == "01":

            report_name = "Reporte Decenal"
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            to_enganches = 0
            to_men = 0
            to_men_wf = 0
            to_dev = 0
            to_otr_prod = 0
            pay = self.env['account.payment']
            move = self.env['account.move']
            busqueda = [('type', '=', 'out_invoice'), ('state', '=', 'posted'), ('categoria_producto', '=', Categorias.id), ('invoice_payment_state','!=','paid')]
            bus_prov = []
            sel_last_p = "-"
            sel_last_c = "-"

            total1 = 0
            total2 = 0
            total3 = "-"
            total4 = "-"

            if str(today.strftime("%d")) == "01":

                bus_prov = [('type', '=', 'in_invoice'), ('state', '=', 'posted'),('invoice_payment_state', '!=', 'paid')]

                one_mes=fields.Date.from_string(today)
                one = datetime.strptime(str(one_mes.strftime("%Y-%m-")) + "01", "%Y-%m-%d")
                one -= relativedelta(days=1)

                last_mes -= relativedelta(months=1)
                ms = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "21", "%Y-%m-%d")
                datei = fields.Date.from_string(ms)
                datef = one

                # ms = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "21", "%Y-%m-%d")
                # mf = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "26", "%Y-%m-%d")

                # datei = fields.Date.from_string(ms)
                # datef = fields.Date.from_string(mf)

                searmen = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 8),

                ]
                searmenwf = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 11),

                ]
                searant = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 9),

                ]
                searantwf = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 10),

                ]
                devent = [
                    ('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                    ('codigo_prod.code', '=', '101.01.001.3'),

                ]

                mensualidades = pay.search(searmen)
                mensualidadeswf = pay.search(searmenwf)
                anticipo = pay.search(searant)
                anticipowf = pay.search(searantwf)
                devent = pay.search(devent)
                otros_productos = pay.search(
                    [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                     ('codigo_prod.code', '=', '601.84.01.04'), ])
                mensu = 0
                mora = 0
                mensuwf = 0
                morawf = 0
                antic = 0
                anticwf = 0
                dev = 0
                otr_prod = self.searchdat(otros_productos)

                for men in mensualidades:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensu -= men.amount
                    mensu += men.amount
                    # an = self.env['account.payment'].search([('name', '=',men.name)])
                    mora += float(men.ji_moratorio)

                for men in mensualidadeswf:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensuwf -= men.amount
                    mensuwf += men.amount
                    # an = self.env['account.payment'].search([('name', '=', men.name)])
                    morawf += float(men.ji_moratorio)

                ant_mora = 0
                for ant in anticipo:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            antic -= ant.amount
                    antic += ant.amount
                    ant_mora += float(ant.ji_moratorio)

                ant_mora_wf = 0
                for ant in anticipowf:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            anticwf -= ant.amount
                    anticwf += ant.amount
                    ant_mora_wf += float(ant.ji_moratorio)
                for de in devent:
                    dev += de.amount

                sheet['F7'] = antic + anticwf + ant_mora_wf +ant_mora
                self.ks_apply_left(sheet['F7'], True, True, 12, False,1)
                sheet['F8'] = mensu + mora
                self.ks_apply_left(sheet['F8'], True, True, 12, False,1)
                sheet['F9'] = mensuwf + morawf
                self.ks_apply_left(sheet['F9'], True, True, 12, False,1)
                sheet['F10'] = dev * -1
                self.ks_apply_left(sheet['F10'], True, True, 12, False,2)
                sheet['F11'] = otr_prod
                self.ks_apply_left(sheet['F11'], True, True, 12, False,1)

                pay_prove = [

                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'outbound'),
                    ('categoria_producto', '=', Categorias.id)
                ]

                pay_out = pay.search(pay_prove)

                pay_prov = self.searchdat(pay_out)

                sheet['F14'] = pay_prov - dev
                self.ks_apply_left(sheet['F14'], True, True, 12, False, 1)

                bus_prov.append(('date', '<=', datef))
                bus_prov.append(('categoria_producto', '=', Categorias.id))
                prov_pay = move.search(bus_prov)
                resd_prov = self.searchresidual(prov_pay)
                sheet['F15'] = resd_prov
                self.ks_apply_left(sheet['F15'], True, True, 12, False, 1)

                bus_prov = [('type', '=', 'in_invoice'),
                            ('state', '=', 'posted'), ('date', '>=', datei), ('date', '<=', datef)
                            ]

                prov_all = move.search(bus_prov)
                total2 = (pay_prov) - self.searchrtotal(prov_all)

                busqueda.append(('date', '<=', datef))
                por_pay = move.search(busqueda)
                residual = self.searchresidual(por_pay)
                sheet['F16'] = residual
                self.ks_apply_left(sheet['F16'], True, True, 12, False, 1)

                busqueda = [('type', '=', 'out_invoice'), ('state', '=', 'posted'),
                            ('categoria_producto', '=', Categorias.id),
                            ('date', '>=', datei), ('date', '<=', datef),
                            ('invoice_payment_state', '!=', 'paid')]

                por_payvt = move.search(busqueda)
                residualvb = self.searchresidual(por_payvt)
                total4 = "+ SUM(F7:F11) - " + str(residualvb)

                sel_last_p = "=F15"
                sel_last_c = "=F16"

                sheet['C19'] = len(por_pay)
                self.ks_apply_left(sheet['C19'], False, True, 12, False, 0)
                sheet['D19'] = 1
                self.ks_apply_left(sheet['D19'], False, True, 12, False, 3)
                sheet['E19'] = residual
                self.ks_apply_left(sheet['E19'], True, True, 12, False, 1)

                al_dia, un_mes, dos_mes, mas_mes = 0, 0, 0, 0
                pay0, pay1, pay2, pay3 = 0, 0, 0, 0
                mora, mora1, mora2, mora3 = 0, 0, 0, 0

                for ac in por_pay:
                    if ac.total_mes1 == 0:
                        al_dia += 1
                        pay0 += ac.saldo_pend
                        mora += ac.total_moratorium
                    if ac.total_mes1 == 1:
                        un_mes += 1
                        pay1 += ac.saldo_pend
                        mora1 += ac.total_moratorium
                    if ac.total_mes1 == 2:
                        dos_mes += 1
                        pay2 += ac.saldo_pend
                        mora2 += ac.total_moratorium
                    if ac.total_mes1 > 2:
                        mas_mes += 1
                        pay3 += ac.saldo_pend
                        mora3 += ac.total_moratorium

                sheet['C20'] = al_dia
                self.ks_apply_left(sheet['C20'], False, True, 12, False, 0)
                sheet['D20'] = '=C20/$C$19'
                self.ks_apply_left(sheet['D20'], False, True, 12, False, 3)
                sheet['E20'] = pay0
                self.ks_apply_left(sheet['E20'], True, True, 12, False, 1)
                sheet['F20'] = mora
                self.ks_apply_left(sheet['F20'], True, True, 12, False, 1)

                sheet['C21'] = un_mes
                self.ks_apply_left(sheet['C21'], False, True, 12, False, 0)
                sheet['D21'] = '=C21/$C$19'
                self.ks_apply_left(sheet['D21'], False, True, 12, False, 3)
                sheet['E21'] = pay1
                self.ks_apply_left(sheet['E21'], True, True, 12, False, 1)
                sheet['F21'] = mora1
                self.ks_apply_left(sheet['F21'], True, True, 12, False, 1)

                sheet['C22'] = dos_mes
                self.ks_apply_left(sheet['C22'], False, True, 12, False, 0)
                sheet['D22'] = '=C22/$C$19'
                self.ks_apply_left(sheet['D22'], False, True, 12, False, 3)
                sheet['E22'] = pay2
                self.ks_apply_left(sheet['E22'], True, True, 12, False, 1)
                sheet['F22'] = mora2
                self.ks_apply_left(sheet['F22'], True, True, 12, False, 1)

                sheet['C23'] = mas_mes
                self.ks_apply_left(sheet['C23'], False, True, 12, False, 0)
                sheet['D23'] = '=C23/$C$19'
                self.ks_apply_left(sheet['D23'], False, True, 12, False, 3)
                sheet['E23'] = pay3
                self.ks_apply_left(sheet['E23'], True, True, 12, False, 1)
                sheet['F23'] = mora3
                self.ks_apply_left(sheet['F23'], True, True, 12, False, 1)

            if str(today.strftime("%d")) >= "21" or str(last_mes.strftime("%d")) == "01":

                bus_prov = [('type', '=', 'in_invoice'), ('state', '=', 'posted'),
                            ('invoice_payment_state', '!=', 'paid')]

                ms = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "11", "%Y-%m-%d")
                mf = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "20", "%Y-%m-%d")


                datei = fields.Date.from_string(ms)
                datef = fields.Date.from_string(mf)


                searmen = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 8),

                ]
                searmenwf = [

                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 11),

                ]
                searant = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 9),

                ]
                searantwf = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 10),

                ]
                devent = [
                    ('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                    ('codigo_prod.code', '=', '101.01.001.3'),

                ]

                mensualidades = pay.search(searmen)
                mensualidadeswf = pay.search(searmenwf)
                anticipo = pay.search(searant)
                anticipowf = pay.search(searantwf)
                devent = pay.search(devent)
                otros_productos = pay.search(
                    [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                     ('codigo_prod.code', '=', '601.84.01.04'), ])
                mensu = 0
                mora = 0
                mensuwf = 0
                morawf = 0
                antic = 0
                anticwf = 0
                dev = 0
                otr_prod = self.searchdat(otros_productos)

                for men in mensualidades:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensu -= men.amount
                    mensu += men.amount
                    # an = self.env['account.payment'].search([('name', '=',men.name)])
                    mora += float(men.ji_moratorio)

                for men in mensualidadeswf:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensuwf -= men.amount
                    mensuwf += men.amount
                    # an = self.env['account.payment'].search([('name', '=', men.name)])
                    morawf += float(men.ji_moratorio)

                ant_mora = 0
                for ant in anticipo:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            antic -= ant.amount
                    antic += ant.amount
                    ant_mora += float(ant.ji_moratorio)

                ant_mora_wf = 0
                for ant in anticipowf:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            anticwf -= ant.amount
                    anticwf += ant.amount
                    ant_mora_wf += float(ant.ji_moratorio)
                for de in devent:
                    dev += de.amount

                sheet['E7'] = antic + anticwf + ant_mora_wf + ant_mora
                self.ks_apply_left(sheet['E7'], True, True, 12, False, 1)
                sheet['E8'] = mensu + mora
                self.ks_apply_left(sheet['E8'], True, True, 12, False, 1)
                sheet['E9'] = mensuwf + morawf
                self.ks_apply_left(sheet['E9'], True, True, 12, False, 1)
                sheet['E10'] = dev * -1
                self.ks_apply_left(sheet['E10'], True, True, 12, False, 2)
                sheet['E11'] = otr_prod
                self.ks_apply_left(sheet['E11'], True, True, 12, False, 1)

                pay_prove = [

                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'outbound'),
                    ('categoria_producto', '=', Categorias.id)
                ]

                pay_out = pay.search(pay_prove)

                pay_prov = self.searchdat(pay_out)

                sheet['E14'] = pay_prov - dev
                self.ks_apply_left(sheet['E14'], True, True, 12, False, 1)

                bus_prov.append(('date', '<=', datef))
                bus_prov.append(('categoria_producto', '=', Categorias.id))
                prov_pay = move.search(bus_prov)
                resd_prov = self.searchresidual(prov_pay)
                # sheet['E15'] = resd_prov
                if total2 == 0:
                    sheet['E15'] = resd_prov
                else:
                    sheet['E15'] = "=F15+" + str(total2)
                self.ks_apply_left(sheet['E15'], True, True, 12, False, 1)

                busqueda.append(('date', '<=', datef))
                por_pay = move.search(busqueda)
                residual = self.searchresidual(por_pay)
                if total4 == "-":
                    sheet['E16'] = residual
                else:
                    sheet['E16'] = "=F16 " +str(total4)

                self.ks_apply_left(sheet['E16'], True, True, 12, False, 1)

                bus_prov = [('type', '=', 'in_invoice'),
                            ('state', '=', 'posted'), ('date', '>=', datei), ('date', '<=', datef)
                            ]

                prov_all = move.search(bus_prov)
                total1 = (pay_prov) - self.searchrtotal(prov_all)

                busqueda = [('type', '=', 'out_invoice'), ('state', '=', 'posted'),
                            ('categoria_producto', '=', Categorias.id),
                            ('date', '>=', datei), ('date', '<=', datef),
                            ('invoice_payment_state', '!=', 'paid')]

                por_payvt = move.search(busqueda)
                residualvb = self.searchresidual(por_payvt)
                total3 = "+ SUM(E7:E11) - " + str(residualvb)


                if sel_last_p =="-":
                    sel_last_p = "=E15"
                    sel_last_c = "=E16"

                    sheet['C19'] = len(por_pay)
                    self.ks_apply_left(sheet['C19'], False, True, 12, False, 0)
                    sheet['D19'] = 1
                    self.ks_apply_left(sheet['D19'], False, True, 12, False, 3)
                    sheet['E19'] = residual
                    self.ks_apply_left(sheet['E19'], True, True, 12, False, 1)

                    al_dia, un_mes, dos_mes, mas_mes = 0, 0, 0, 0
                    pay0, pay1, pay2, pay3 = 0, 0, 0, 0
                    mora, mora1, mora2, mora3 = 0, 0, 0, 0


                    for ac in por_pay:
                        if ac.total_mes1 == 0:
                            al_dia += 1
                            pay0 += ac.saldo_pend
                            mora += ac.total_moratorium
                        if ac.total_mes1 == 1:
                            un_mes += 1
                            pay1 += ac.saldo_pend
                            mora1 += ac.total_moratorium
                        if ac.total_mes1 == 2:
                            dos_mes += 1
                            pay2 += ac.saldo_pend
                            mora2 += ac.total_moratorium
                        if ac.total_mes1 > 2:
                            mas_mes += 1
                            pay3 += ac.saldo_pend
                            mora3 += ac.total_moratorium



                    sheet['C20'] = al_dia
                    self.ks_apply_left(sheet['C20'], False, True, 12, False, 0)
                    sheet['D20'] = '=C20/$C$19'
                    self.ks_apply_left(sheet['D20'], False, True, 12, False, 3)
                    sheet['E20'] = pay0
                    self.ks_apply_left(sheet['E20'], True, True, 12, False, 1)
                    sheet['F20'] = mora
                    self.ks_apply_left(sheet['F20'], True, True, 12, False, 1)

                    sheet['C21'] = un_mes
                    self.ks_apply_left(sheet['C21'], False, True, 12, False, 0)
                    sheet['D21'] = '=C21/$C$19'
                    self.ks_apply_left(sheet['D21'], False, True, 12, False, 3)
                    sheet['E21'] = pay1
                    self.ks_apply_left(sheet['E21'], True, True, 12, False, 1)
                    sheet['F21'] = mora1
                    self.ks_apply_left(sheet['F21'], True, True, 12, False, 1)


                    sheet['C22'] = dos_mes
                    self.ks_apply_left(sheet['C22'], False, True, 12, False, 0)
                    sheet['D22'] = '=C22/$C$19'
                    self.ks_apply_left(sheet['D22'], False, True, 12, False, 3)
                    sheet['E22'] = pay2
                    self.ks_apply_left(sheet['E22'], True, True, 12, False, 1)
                    sheet['F22'] = mora2
                    self.ks_apply_left(sheet['F22'], True, True, 12, False, 1)



                    sheet['C23'] = mas_mes
                    self.ks_apply_left(sheet['C23'], False, True, 12, False, 0)
                    sheet['D23'] = '=C23/$C$19'
                    self.ks_apply_left(sheet['D23'], False, True, 12, False, 3)
                    sheet['E23'] = pay3
                    self.ks_apply_left(sheet['E23'], True, True, 12, False, 1)
                    sheet['F23'] = mora3
                    self.ks_apply_left(sheet['F23'], True, True, 12, False, 1)

            if str(today.strftime("%d")) >= "11" or str(last_mes.strftime("%d")) == "01" :

                bus_prov = [('type', '=', 'in_invoice'), ('state', '=', 'posted'),
                            ('invoice_payment_state', '!=', 'paid')]

                ms = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "01", "%Y-%m-%d")
                mf = datetime.strptime(str(last_mes.strftime("%Y-%m-")) + "10", "%Y-%m-%d")
                pay = self.env['account.payment']
                datei = fields.Date.from_string(ms)
                datef = fields.Date.from_string(mf)
                searmen = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 8),

                ]
                searmenwf = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 11),

                ]
                searant = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 9),

                ]
                searantwf = [
                    ('categoria_producto', '=', Categorias.id),
                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'inbound'),
                    ('journal_id', '=', 10),

                ]
                devent = [
                    ('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                    ('codigo_prod.code', '=', '101.01.001.3'),

                ]



                mensualidades = pay.search(searmen)
                mensualidadeswf = pay.search(searmenwf)
                anticipo = pay.search(searant)
                anticipowf = pay.search(searantwf)
                devent = pay.search(devent)
                otros_productos = pay.search(
                    [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                     ('codigo_prod.code', '=', '601.84.01.04'), ])
                mensu = 0
                mora = 0
                mensuwf = 0
                morawf = 0
                antic = 0
                anticwf = 0
                dev = 0
                otr_prod = self.searchdat(otros_productos)

                for men in mensualidades:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensu -= men.amount
                    mensu += men.amount
                    # an = self.env['account.payment'].search([('name', '=',men.name)])
                    mora += float(men.ji_moratorio)

                for men in mensualidadeswf:
                    if men.codigo_prod:
                        if men.codigo_prod.code == '601.84.01.04':
                            mensuwf -= men.amount
                    mensuwf += men.amount
                    # an = self.env['account.payment'].search([('name', '=', men.name)])
                    morawf += float(men.ji_moratorio)

                ant_mora = 0
                for ant in anticipo:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            antic -= ant.amount
                    antic += ant.amount
                    ant_mora += float(ant.ji_moratorio)

                ant_mora_wf = 0
                for ant in anticipowf:
                    if ant.codigo_prod:
                        if ant.codigo_prod.code == '601.84.01.04':
                            anticwf -= ant.amount
                    anticwf += ant.amount
                    ant_mora_wf += float(ant.ji_moratorio)
                for de in devent:
                    dev += de.amount

                sheet['D7'] = antic + anticwf + ant_mora_wf + ant_mora
                self.ks_apply_left(sheet['D7'], True, True, 12, False, 1)
                sheet['D8'] = mensu + mora
                self.ks_apply_left(sheet['D8'], True, True, 12, False, 1)
                sheet['D9'] = mensuwf + morawf
                self.ks_apply_left(sheet['D9'], True, True, 12, False, 1)
                sheet['D10'] = dev * -1
                self.ks_apply_left(sheet['D10'], True, True, 12, False, 2)
                sheet['D11'] = otr_prod
                self.ks_apply_left(sheet['D11'], True, True, 12, False, 1)

                pay_prove = [

                    ('state', '=', 'posted'),
                    ('payment_date', '>=', datei),
                    ('payment_date', '<=', datef),
                    ('payment_type', '=', 'outbound'),
                    ('payment_type', '=', 'outbound'),
                    ('categoria_producto', '=', Categorias.id)
                ]

                pay_out = pay.search(pay_prove)

                pay_prov = self.searchdat(pay_out)

                sheet['D14'] = pay_prov - dev
                self.ks_apply_left(sheet['D14'], True, True, 12, False, 1)

                bus_prov.append(('date', '<=', datef))
                bus_prov.append(('categoria_producto', '=', Categorias.id))
                prov_pay = move.search(bus_prov)
                resd_prov = self.searchresidual(prov_pay)
                if total1 == 0:
                    sheet['D15'] = resd_prov
                else:
                    sheet['D15'] = "=E15+" + str(total1 + total2)
                self.ks_apply_left(sheet['D15'], True, True, 12, False, 1)

                busqueda.append(('date', '<=', datef))
                por_pay = move.search(busqueda)
                residual = self.searchresidual(por_pay)
                if total3 == "-":
                    sheet['D16'] = residual
                else:
                    sheet['D16'] = "=E16 " +str(total3)
                self.ks_apply_left(sheet['D16'], True, True, 12, False, 1)

                if sel_last_p =="-":
                    sel_last_p = "=D15"
                    sel_last_c = "=D16"

                    sheet['C19'] = len(por_pay)
                    self.ks_apply_left(sheet['C19'], False, True, 12, False, 0)
                    sheet['D19'] = 1
                    self.ks_apply_left(sheet['D19'], False, True, 12, False, 3)
                    sheet['E19'] = residual
                    self.ks_apply_left(sheet['E19'], True, True, 12, False, 1)

                    al_dia, un_mes, dos_mes, mas_mes = 0, 0, 0, 0
                    pay0, pay1, pay2, pay3 = 0, 0, 0, 0
                    mora, mora1, mora2, mora3 = 0, 0, 0, 0

                    for ac in por_pay:
                        if ac.total_mes1 == 0:
                            al_dia += 1
                            pay0 += ac.saldo_pend
                            mora += ac.total_moratorium
                        if ac.total_mes1 == 1:
                            un_mes += 1
                            pay1 += ac.saldo_pend
                            mora1 += ac.total_moratorium
                        if ac.total_mes1 == 2:
                            dos_mes += 1
                            pay2 += ac.saldo_pend
                            mora2 += ac.total_moratorium
                        if ac.total_mes1 > 2:
                            mas_mes += 1
                            pay3 += ac.saldo_pend
                            mora3 += ac.total_moratorium

                    sheet['C20'] = al_dia
                    self.ks_apply_left(sheet['C20'], False, True, 12, False, 0)
                    sheet['D20'] = '=C20/$C$19'
                    self.ks_apply_left(sheet['D20'], False, True, 12, False, 3)
                    sheet['E20'] = pay0
                    self.ks_apply_left(sheet['E20'], True, True, 12, False, 1)
                    sheet['F20'] = mora
                    self.ks_apply_left(sheet['F20'], True, True, 12, False, 1)

                    sheet['C21'] = un_mes
                    self.ks_apply_left(sheet['C21'], False, True, 12, False, 0)
                    sheet['D21'] = '=C21/$C$19'
                    self.ks_apply_left(sheet['D21'], False, True, 12, False, 3)
                    sheet['E21'] = pay1
                    self.ks_apply_left(sheet['E21'], True, True, 12, False, 1)
                    sheet['F21'] = mora1
                    self.ks_apply_left(sheet['F21'], True, True, 12, False, 1)

                    sheet['C22'] = dos_mes
                    self.ks_apply_left(sheet['C22'], False, True, 12, False, 0)
                    sheet['D22'] = '=C22/$C$19'
                    self.ks_apply_left(sheet['D22'], False, True, 12, False, 3)
                    sheet['E22'] = pay2
                    self.ks_apply_left(sheet['E22'], True, True, 12, False, 1)
                    sheet['F22'] = mora2
                    self.ks_apply_left(sheet['F22'], True, True, 12, False, 1)

                    sheet['C23'] = mas_mes
                    self.ks_apply_left(sheet['C23'], False, True, 12, False, 0)
                    sheet['D23'] = '=C23/$C$19'
                    self.ks_apply_left(sheet['D23'], False, True, 12, False, 3)
                    sheet['E23'] = pay3
                    self.ks_apply_left(sheet['E23'], True, True, 12, False, 1)
                    sheet['F23'] = mora3
                    self.ks_apply_left(sheet['F23'], True, True, 12, False, 1)



            sheet['C7'] = "=SUM(D7:F7)"
            self.ks_apply_left(sheet['C7'], True, True, 12, False, 1)
            sheet['C8'] = "=SUM(D8:F8)"
            self.ks_apply_left(sheet['C8'], True, True, 12, False, 1)
            sheet['C9'] = "=SUM(D9:F9)"
            self.ks_apply_left(sheet['C9'], True, True, 12, False, 1)
            sheet['C10'] = "=SUM(D10:F10)"
            self.ks_apply_left(sheet['C10'], True, True, 12, False, 2)
            sheet['C11'] = "=SUM(D11:F11)"
            self.ks_apply_left(sheet['C11'], True, True, 12, False, 1)

            sheet['C14'] = "=SUM(D14:F14)"
            self.ks_apply_left(sheet['C14'], True, True, 12, False, 1)
            sheet['C15'] = sel_last_p
            self.ks_apply_left(sheet['C15'], True, True, 12, False, 1)
            sheet['C16'] = sel_last_c
            self.ks_apply_left(sheet['C16'], True, True, 12, False, 1)

            one_mes = fields.Date.from_string(today)
            one_mes -= relativedelta(days=1)

            dia = str(one_mes.strftime("%d")) + " de " + self.mes_c(one_mes.strftime("%m")) + " de " + str(
                one_mes.strftime("%Y"))
            self.docenal_create_workbook_header(report_name, sheet, dia)

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            # with NamedTemporaryFile() as tmp:
            #     for column_cells in ws.columns:
            #         length = max(len(self.as_text(cell.value)) for cell in column_cells)
            #         sheet.column_dimensions[column_cells[0].column_letter].width = length
            #     workbook.save(tmp.name)
            #     output = tmp.read()
            #     return output

            output = StringIO()

            filename = ('/tmp/' + str(report_name) + '.xlsx')
            workbook.save(filename)
            fp = open(filename, "rb")
            file_data = fp.read()
            data_record = base64.b64encode(file_data)
            att = {
                'name': str(report_name) + '.xlsx',
                'type': 'binary',
                'datas': data_record,
                'store_fname': data_record,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            }
            template_id = self.env['ir.model.data'].get_object_reference('jibaritolotes', 'mail_template_Docenal')[1]
            email_template_obj = self.env['mail.template'].browse(template_id)
            if template_id:
                data_id = self.env['ir.attachment'].create(att)

                values = email_template_obj.generate_email(325, fields=None)
                values['email_from'] = "jibarito2005@gmail.com"
                # value['email_to'] = partner.email
                # values['email_to'] = str([partner.email for partner in noti.partner_ids]).replace('[', '').replace(']',"'", "")
                values['subject'] = "Reporte Decenal " + dia
                # values['email_cc'] = "jesus.nazareth@ogum.com.mx"
                values['email_to'] = "jibarito2005@gmail.com"
                # raise UserError(_(value['email_to']))
                values['attachment_ids'] = [(6, 0, [data_id.id])]
                mail_mail_obj = self.env['mail.mail']
                print('\n\n\n', values, '\n\n\n')
                msg_id = mail_mail_obj.sudo().create(values)
                if msg_id:
                    mail_mail_obj.sudo().send([msg_id])




    def searchdat(self, purch):
        dev = 0.0
        for de in purch:
            dev += de.amount
        return dev
    def searchmoras(self, purch):
        mora = 0.0
        pago = 0.0
        for de in purch:
            mora += de.total_moratorium
            pago += de.saldo_pend
        return mora, pago
    def searchresidual(self, purch):
        dev = 0.0
        for de in purch:
            dev += de.amount_residual
        return dev
    def searchrtotal(self, purch):
        dev = 0.0
        for de in purch:
            dev += de.amount_total
        return dev
    def set_report_gasto(self, categoria, fechai, fechaf):

            Categorias = categoria
            datef = fechaf
            datei = fechai
            pay = self.env['account.payment']
            account = self.env['account.move.line']
            # account.categoria_venta()
            searmen = [
                ('categoria_producto', '=', Categorias.id),
                ('state', '=', 'posted'),
                ('payment_date', '>=', datei),
                ('payment_date', '<=', datef),
                ('payment_type', '=', 'inbound'),
                ('journal_id', '=', 8),

            ]
            searmenwf = [
                ('categoria_producto', '=', Categorias.id),
                ('state', '=', 'posted'),
                ('payment_date', '>=', datei),
                ('payment_date', '<=', datef),
                ('payment_type', '=', 'inbound'),
                ('journal_id', '=', 11),

            ]
            searant = [
                ('categoria_producto', '=', Categorias.id),
                ('state', '=', 'posted'),
                ('payment_date', '>=', datei),
                ('payment_date', '<=', datef),
                ('payment_type', '=', 'inbound'),
                ('journal_id', '=', 9),

            ]
            searantwf = [
                ('categoria_producto', '=', Categorias.id),
                ('state', '=', 'posted'),
                ('payment_date', '>=', datei),
                ('payment_date', '<=', datef),
                ('payment_type', '=', 'inbound'),
                ('journal_id', '=', 10),

            ]
            devent = [
                ('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                ('codigo_prod.code', '=', '101.01.001.3'),

            ]

            mensualidades = pay.search(searmen)
            mensualidadeswf = pay.search(searmenwf)
            anticipo = pay.search(searant)
            anticipowf = pay.search(searantwf)
            devent = pay.search(devent)

            mejor = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                                ('codigo_prod.code', '=', '501.01.01.01'), ])
            seguridad = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '501.01.01.02'), ])
            levantamientos = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '501.01.01.03'), ])
            laboratorio = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '501.01.01.04'), ])
            jardineria = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '501.01.01.05'), ])
            topografia = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '501.01.01.06'), ])
            otros_gastos = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '601.84.01.01'), ])
            plan_maestro = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '601.84.01.02'), ])
            obra = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                               ('codigo_prod.code', '=', '601.84.01.03'), ])
            gastos_venta = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.01.01'), ])
            promocion_publicidad = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.01.02'), ])
            comisiones = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.01.03'), ])
            predial = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.01.04'), ])
            nomina = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                                 ('codigo_prod.code', '=', '602.02.02.01'), ])
            viaticos = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.02'), ])
            dividendos = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.03'), ])
            transporte = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.04'), ])
            mensajeria = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.05'), ])
            rentas = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                                 ('codigo_prod.code', '=', '602.02.02.06'), ])
            papeleria = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.07'), ])
            servicios_oficina = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.08'), ])
            telefonia_internet = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.09'), ])
            gastos_varios = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.10'), ])
            mantenimiento = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.02.11'), ])
            honorarios = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.03.01'), ])
            juzgado = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.03.02'), ])
            apoyos = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                                 ('codigo_prod.code', '=', '602.02.03.03'), ])
            gastos_legales = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.03.04'), ])
            gastos_legaleswf = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '602.02.03.05'), ])
            indivi = pay.search([('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                                 ('codigo_prod.code', '=', '602.02.03.06'), ])
            otros_productos = pay.search(
                [('state', '=', 'posted'), ('payment_date', '>=', datei), ('payment_date', '<=', datef),
                 ('codigo_prod.code', '=', '601.84.01.04'), ])

            mensu = 0
            mora = 0
            mensuwf = 0
            morawf = 0
            antic = 0
            anticwf = 0
            dev = 0
            ''' Mensualidades y moratorios
                incluyen wf
            '''

            for men in mensualidades:
                if men.codigo_prod:
                    if men.codigo_prod.code == '601.84.01.04':
                        mensu -= men.amount
                mensu += men.amount
                # an = self.env['account.payment'].search([('name', '=',men.name)])
                mora += float(men.ji_moratorio)

            for men in mensualidadeswf:
                if men.codigo_prod:
                    if men.codigo_prod.code == '601.84.01.04':
                        mensuwf -= men.amount
                mensuwf += men.amount
                # an = self.env['account.payment'].search([('name', '=', men.name)])
                morawf += float(men.ji_moratorio)

            ''' Anticipos
             incluyen wf
            '''
            ant_mora = 0
            for ant in anticipo:
                if ant.codigo_prod:
                    if ant.codigo_prod.code == '601.84.01.04':
                        antic -= ant.amount
                antic += ant.amount
                ant_mora += float(ant.ji_moratorio)

            ant_mora_wf = 0

            for ant in anticipowf:
                if ant.codigo_prod:
                    if ant.codigo_prod.code == '601.84.01.04':
                        anticwf -= ant.amount
                anticwf += ant.amount
                ant_mora_wf += float(ant.ji_moratorio)

            for de in devent:
                dev += de.amount
            mej = self.searchdat(mejor)
            seg = self.searchdat(seguridad)
            lev = self.searchdat(levantamientos)
            lab = self.searchdat(laboratorio)
            jar = self.searchdat(jardineria)
            topo = self.searchdat(topografia)
            otr_gast = self.searchdat(otros_gastos)
            plan_maes = self.searchdat(plan_maestro)
            obra = self.searchdat(obra)
            Gast_vent = self.searchdat(gastos_venta)
            prom_pub = self.searchdat(promocion_publicidad)
            comis = self.searchdat(comisiones)
            predi = self.searchdat(predial)
            nomina = self.searchdat(nomina)
            viati = self.searchdat(viaticos)
            divid = self.searchdat(dividendos)
            trans = self.searchdat(transporte)
            mensj = self.searchdat(mensajeria)
            rentas = self.searchdat(rentas)
            papel = self.searchdat(papeleria)
            serv_ofic = self.searchdat(servicios_oficina)
            tel_inter = self.searchdat(telefonia_internet)
            gast_var = self.searchdat(gastos_varios)
            mant = self.searchdat(mantenimiento)

            honor = self.searchdat(honorarios)
            juzg = self.searchdat(juzgado)
            apoyos = self.searchdat(apoyos)
            gast_leg = self.searchdat(gastos_legales)
            gast_legwf = self.searchdat(gastos_legaleswf)
            indivi = self.searchdat(indivi)
            otr_prod = self.searchdat(otros_productos)

            vent_net = mensu + mensuwf + mora + morawf + antic + anticwf + ant_mora + ant_mora_wf - dev
            cost_vent = mej + seg + lev + lab + jar + topo
            util_perd_bru = vent_net - cost_vent
            Gast_ventg = Gast_vent + prom_pub + comis + predi
            gast_admin = nomina + trans + mensj + rentas + papel + serv_ofic + tel_inter + gast_var + mant
            gas_cons = viati + divid
            gas_leg = honor + juzg + apoyos + gast_leg + gast_legwf + indivi
            util_perd_oper = util_perd_bru - Gast_ventg - gast_admin - gas_cons - gas_leg
            otrs_gast = otr_gast + plan_maes + obra
            otrs_prod = otr_prod
            util_perd_ejer = util_perd_oper - otrs_gast + otrs_prod

            egre = cost_vent + Gast_ventg + gast_admin + gas_cons + gas_leg + otrs_gast
            por = 0
            data = {
                'categoria': Categorias.name,
                'por': por,
                'datei': datei.strftime('%d-%m-%Y'),
                'datef': datef.strftime('%d-%m-%Y'),
                'mens': "$ {0:,.2f}".format(mensu),
                'menswf': "$ {0:,.2f}".format(mensuwf),
                'mora': "$ {0:,.2f}".format(mora),
                'morawf': "$ {0:,.2f}".format(morawf),
                'anticipo': "$ {0:,.2f}".format(antic),
                'anticipowf': "$ {0:,.2f}".format(anticwf),
                'ant_mora': "$ {0:,.2f}".format(ant_mora),
                'ant_mora_wf': "$ {0:,.2f}".format(ant_mora_wf),
                'dev': "$ {0:,.2f}".format(dev),
                'ven_net': "$ {0:,.2f}".format(vent_net),
                'ven_net_o': "$ {0:,.2f}".format(vent_net + otr_prod),
                'mejoras': "$ {0:,.2f}".format(mej),
                'seguridad': "$ {0:,.2f}".format(seg),
                'levant': "$ {0:,.2f}".format(lev),
                'labora': "$ {0:,.2f}".format(lab),
                'jardin': "$ {0:,.2f}".format(jar),
                'topog': "$ {0:,.2f}".format(topo),
                'cost_vent': "$ {0:,.2f}".format(cost_vent),
                'util_perd_bru': "$ {0:,.2f}".format(util_perd_bru),
                'Gast_vent': "$ {0:,.2f}".format(Gast_vent),
                'prom_pub': "$ {0:,.2f}".format(prom_pub),
                'comis': "$ {0:,.2f}".format(comis),
                'predi': "$ {0:,.2f}".format(predi),
                'Gast_ventg': "$ {0:,.2f}".format(Gast_ventg),
                'gast_admin': "$ {0:,.2f}".format(gast_admin),
                'nomina': "$ {0:,.2f}".format(nomina),
                'trans': "$ {0:,.2f}".format(trans),
                'mensj': "$ {0:,.2f}".format(mensj),
                'rentas': "$ {0:,.2f}".format(rentas),
                'papel': "$ {0:,.2f}".format(papel),
                'serv_ofic': "$ {0:,.2f}".format(serv_ofic),
                'tel_inter': "$ {0:,.2f}".format(tel_inter),
                'gast_var': "$ {0:,.2f}".format(gast_var),
                'mant': "$ {0:,.2f}".format(mant),
                'gas_cons': "$ {0:,.2f}".format(gas_cons),
                'viati': "$ {0:,.2f}".format(viati),
                'divid': "$ {0:,.2f}".format(divid),
                'gas_leg': "$ {0:,.2f}".format(gas_leg),
                'honor': "$ {0:,.2f}".format(honor),
                'juzg': "$ {0:,.2f}".format(juzg),
                'apoyos': "$ {0:,.2f}".format(apoyos),
                'gast_leg': "$ {0:,.2f}".format(gast_leg),
                'gast_legwf': "$ {0:,.2f}".format(gast_legwf),
                'indivi': "$ {0:,.2f}".format(indivi),
                'util_perd_oper': "$ {0:,.2f}".format(util_perd_oper),
                'otrs_gast': "$ {0:,.2f}".format(otrs_gast),
                'otr_gast': "$ {0:,.2f}".format(otr_gast),
                'plan_maes': "$ {0:,.2f}".format(plan_maes),
                'obra': "$ {0:,.2f}".format(obra),
                'otrs_prod': "$ {0:,.2f}".format(otrs_prod),
                'otr_prod': "$ {0:,.2f}".format(otr_prod),
                'util_perd_ejer': "$ {0:,.2f}".format(util_perd_ejer),
                'egre': "$ {0:,.2f}".format(egre),
                'resul': "$ {0:,.2f}".format(vent_net + otrs_prod - egre),
               'dif': "$ {0:,.2f}".format(vent_net - egre - util_perd_ejer + otrs_prod),

            }
            return self.env.ref('jibaritolotes.report_gastosv2').render_qweb_pdf(self, data=data)
            # return self.env.ref('jibaritolotes.report_gastos').report_action(self, data=data)

